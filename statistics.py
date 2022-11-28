import csv

from line_profiler import LineProfiler
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit
import datetime
from dateutil.parser import parse

profiler = LineProfiler()


class Helper:
    @staticmethod
    def parse_year_from_date_slice(date: str) -> int:
        return int(date[:4])

    @staticmethod
    def parse_year_from_date_datetime(date: str) -> int:
        return datetime.datetime.strptime(date, '%Y-%m-%dT%H:%M:%S%z').year

    @staticmethod
    def parse_year_from_date_dateutil(date: str) -> int:
        return parse(date).year


class Vacancy:
    """
    Класс для представления вакансии

    Attributes:
        name (str): Название
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
        salary_average (float): Средняя зарплата
        area_name (str): Место, где есть вакансия
        year (int): Год создания вакансии
    """
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    @profile
    def __init__(self, vacancy: dict[str, str]) -> None:
        """
        Инициализирует объект Vacancy, выполняет конвертацию для некоторых полей, считает среднюю зарплату

        :param vacancy: Словарь вакансии вида [str, str]
        :returns: None
        """
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = Helper.parse_year_from_date_slice(vacancy['published_at'])
        # self.year = Helper.parse_year_from_date_datetime(vacancy['published_at'])
        # self.year = Helper.parse_year_from_date_dateutil(vacancy['published_at'])


class DataSet:
    """
    Класс для представления данных

    Attributes:
        file_name (str): Название файла
        vacancy_name (str): Название вакансии
    """

    @profile
    def __init__(self, file: str, vacancy: str) -> None:
        """
        Инициализирует объект Dataset

        :param file: Название файла
        :param vacancy: Название вакансии

        >>> type(DataSet("123.csv", "Программист")).__name__
        'DataSet'
        >>> type(DataSet("123.csv", "Программист").file_name).__name__
        'str'
        >>> type(DataSet("123.csv", "Программист").vacancy_name).__name__
        'str'
        """
        self.file_name = file
        self.vacancy_name = vacancy

    @staticmethod
    @profile
    def increment(subject: dict, key, value) -> None:
        """
        Если в subject есть значение с ключом key: увеличивает его на value, иначе: присваивает ему значение value

        :param subject: Словарь объектов
        :param key: Ключ для поиска элемента
        :param value: Значение для инкремента или присваивания
        :return: None
        """
        if key in subject:
            subject[key] += value
        else:
            subject[key] = value

    @staticmethod
    @profile
    def get_average_dict(data: dict) -> dict:
        """
        Создаёт новый словарь из данного, где элементы - среднее значение

        :param data: Словарь с данными
        :return: Массив средних

        >>> DataSet.get_average_dict({1: [2, 5], 2: [3, 6]})
        {1: 3, 2: 4}
        """
        result = {}
        for key, data in data.items():
            result[key] = int(sum(data) / len(data))
        return result

    @profile
    def csv_reader(self) -> dict:
        """
        Открывает файл и лениво возвращает словари с данными вакансии
        """
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            titles = next(reader)
            titles_count = len(titles)
            for row in reader:
                if '' not in row and len(row) == titles_count:
                    yield dict(zip(titles, row))

    @profile
    def get_statistics(self) -> (dict, dict, dict, dict, dict, dict):
        """
        Формирует статистику по вакансиям и возвращает кортеж с данными

        :returns (stat_salary, vacancies_number, stat_salary_by_vac, vacs_per_name, stat_salary_by_city,
         top_salary_by_year): Статистика по зп, статистика по числу вакансий, статистика вакансий по ЗП,
         статистика вакансий по названию, статистика вакансий по городам
        """
        salary = {}
        salary_of_vacancy_name = {}
        salary_city = {}
        vacancies_count = 0
        for vacancy in self.csv_reader():
            vacancy = Vacancy(vacancy)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_of_vacancy_name, vacancy.year, [vacancy.salary_average])
            self.increment(salary_city, vacancy.area_name, [vacancy.salary_average])
            vacancies_count += 1
        vacancies_number = dict([(key, len(value)) for key, value in salary.items()])
        vacs_per_name = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])
        if not salary_of_vacancy_name:
            salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
            vacs_per_name = dict([(key, 0) for key, value in vacancies_number.items()])
        stat_salary = self.get_average_dict(salary)
        stat_salary_by_vac = self.get_average_dict(salary_of_vacancy_name)
        stat_salary_by_city = self.get_average_dict(salary_city)
        stat_salary_by_year = {}
        for year, salaries in salary_city.items():
            stat_salary_by_year[year] = round(len(salaries) / vacancies_count, 4)
        stat_salary_by_year = list(filter(lambda a: a[-1] >= 0.01,
                                          [(key, value) for key, value in stat_salary_by_year.items()]))
        stat_salary_by_year.sort(key=lambda a: a[-1], reverse=True)
        top_salary_by_year = stat_salary_by_year.copy()
        stat_salary_by_year = dict(stat_salary_by_year)
        stat_salary_by_city = list(filter(lambda a: a[0] in list(stat_salary_by_year.keys()),
                                          [(key, value) for key, value in stat_salary_by_city.items()]))
        stat_salary_by_city.sort(key=lambda a: a[-1], reverse=True)
        stat_salary_by_city = dict(stat_salary_by_city[:10])
        top_salary_by_year = dict(top_salary_by_year[:10])
        return stat_salary, vacancies_number, stat_salary_by_vac, vacs_per_name, stat_salary_by_city, top_salary_by_year

    @staticmethod
    @profile
    def print_statistic(salary_by_year: dict, vacs_per_year: dict, salary_by_vac: dict, count_by_vac: dict,
                        salary_by_city: dict, city_percents: dict) -> None:
        """
        Печатает статистику

        :param salary_by_year: Статистика зарплат по годам
        :param vacs_per_year: Статистика количества вакансий по годам
        :param salary_by_vac: Статистика зарплаты по годам для выбранной профессии
        :param count_by_vac: Статистика количества вакансий по годам для выбранной профессии
        :param salary_by_city: Статистика зарплаты по городам
        :param city_percents: Статистика доли вакансий по городам
        :return: None
        """
        print(f'Динамика уровня зарплат по годам: {salary_by_year}')
        print(f'Динамика количества вакансий по годам: {vacs_per_year}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {salary_by_vac}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {count_by_vac}')
        print(f'Уровень зарплат по городам (в порядке убывания): {salary_by_city}')
        print(f'Доля вакансий по городам (в порядке убывания): {city_percents}')


class InputConnect:
    """
    Работает с вводом пользователя, составляет датасет, создаёт файлы отчетов: XLSX, png, pdf

    Attributes:
        file_name (str): Название файла
        vacancy_name (str): Название профессии
    """

    @profile
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')
        dataset = DataSet(self.file_name, self.vacancy_name)
        salary_by_year, vacs_per_year, salary_by_vac, count_by_vac, salary_by_city, city_percents = \
            dataset.get_statistics()
        report = Report(self.vacancy_name, salary_by_year, vacs_per_year, salary_by_vac,
                        count_by_vac, salary_by_city, city_percents)
        report.create_xlsx_file()
        report.generate_image()
        report.generate_pdf()


class Report:
    """
    Генерирует отчеты XLSX,  PNG, PDF

    Attributes:
        wb (Workbook): XLSX файл
        vacancy_name (str): Название вакансии
        salary_by_year (dict): Статистика зарплат по годам
        vacs_per_year (dict): Статистика количества вакансий по годам
        salary_by_vac (dict): Статистика зарплаты по годам для выбранной профессии
        count_by_vac (dict): Статистика количества вакансий по годам для выбранной профессии
        salary_by_city (dict): Статистика зарплаты по городам
        city_percents (dict): Статистика доли вакансий по городам
    """

    @profile
    def __init__(self, vacancy_name: str, salary_by_year: dict, vacs_per_year: dict, salary_by_vac: dict,
                 count_by_vac: dict, salary_by_city: dict, city_percents: dict) -> None:
        """
        Инициализирует объект отчёта и XLSX файл

        :param vacancy_name: Название вакансии
        :param salary_by_year: Статистика зарплат по годам
        :param vacs_per_year: Статистика количества вакансий по годам
        :param salary_by_vac: Статистика зарплаты по годам для выбранной профессии
        :param count_by_vac: Статистика количества вакансий по годам для выбранной профессии
        :param salary_by_city: Статистика зарплаты по городам
        :param city_percents: Статистика доли вакансий по городам
        :return: None
        """
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.salary_by_year = salary_by_year
        self.vacs_per_year = vacs_per_year
        self.salary_by_vac = salary_by_vac
        self.count_by_vac = count_by_vac
        self.salary_by_city = salary_by_city
        self.city_percents = city_percents

    @profile
    def create_xlsx_file(self) -> None:
        """
        Создаёт XLSX файл-отчёт

        :return: None
        """
        year_sheet = self.wb.active
        year_sheet.title = 'Статистика по годам'
        year_sheet.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name,
                           'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.salary_by_year.keys():
            year_sheet.append([year, self.salary_by_year[year], self.salary_by_vac[year],
                               self.vacs_per_year[year], self.count_by_vac[year]])
        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name,
                 ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(widths) > i:
                    if len(cell) > widths[i]:
                        widths[i] = len(cell)
                else:
                    widths += [len(cell)]
        for i, column_width in enumerate(widths, 1):
            year_sheet.column_dimensions[get_column_letter(i)].width = column_width + 2
        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city1, value1), (city2, value2) in zip(self.salary_by_city.items(), self.city_percents.items()):
            data.append([city1, value1, '', city2, value2])
        city_sheet = self.wb.create_sheet('Статистика по городам')
        for row in data:
            city_sheet.append(row)
        widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(widths) > i:
                    if len(cell) > widths[i]:
                        widths[i] = len(cell)
                else:
                    widths += [len(cell)]
        for i, column_width in enumerate(widths, 1):
            city_sheet.column_dimensions[get_column_letter(i)].width = column_width + 2
        font_bold = Font(bold=True)
        for col in 'ABCDE':
            year_sheet[col + '1'].font = font_bold
            city_sheet[col + '1'].font = font_bold
        for index, _ in enumerate(self.salary_by_city):
            city_sheet['E' + str(index + 2)].number_format = '0.00%'
        thin = Side(border_style='thin', color='00000000')
        for row in range(len(data)):
            for col in 'ABDE':
                city_sheet[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)
        for row, _ in enumerate(self.salary_by_year):
            for col in 'ABCDE':
                year_sheet[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)
        self.wb.save(filename='report.xlsx')

    @profile
    def generate_image(self) -> None:
        """
        Генерирует PNG-изображение со статистикой
        :return: None
        """
        fig, ((ax_1, ax_2), (ax_3, ax_4)) = plt.subplots(nrows=2, ncols=2)

        bar1 = ax_1.bar(np.array(list(self.salary_by_year.keys())) - 0.4, self.salary_by_year.values(), width=0.4)
        bar2 = ax_1.bar(np.array(list(self.salary_by_year.keys())), self.salary_by_vac.values(), width=0.4)
        ax_1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax_1.grid(axis='y')
        ax_1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        ax_1.set_xticks(np.array(list(self.salary_by_year.keys())) - 0.2, list(self.salary_by_year.keys()), rotation=90)
        ax_1.xaxis.set_tick_params(labelsize=8)
        ax_1.yaxis.set_tick_params(labelsize=8)

        ax_2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = ax_2.bar(np.array(list(self.vacs_per_year.keys())) - 0.4, self.vacs_per_year.values(), width=0.4)
        bar2 = ax_2.bar(np.array(list(self.vacs_per_year.keys())), self.count_by_vac.values(), width=0.4)
        ax_2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
                    prop={'size': 8})
        ax_2.set_xticks(np.array(list(self.vacs_per_year.keys())) - 0.2, list(self.vacs_per_year.keys()), rotation=90)
        ax_2.grid(axis='y')
        ax_2.xaxis.set_tick_params(labelsize=8)
        ax_2.yaxis.set_tick_params(labelsize=8)

        ax_3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax_3.barh(list([str(a)
                       .replace(' ', '\n')
                       .replace('-', '-\n') for a in reversed(list(self.salary_by_city.keys()))]),
                  list(reversed(list(self.salary_by_city.values()))), color='blue', height=0.5, align='center')
        ax_3.yaxis.set_tick_params(labelsize=6)
        ax_3.xaxis.set_tick_params(labelsize=8)
        ax_3.grid(axis='x')

        ax_4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.city_percents.values()])
        ax_4.pie(list(self.city_percents.values()) + [other], labels=list(self.city_percents.keys()) + ['Другие'],
                 textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')

    @profile
    def generate_pdf(self) -> None:
        """
        Генерирует PDF-файл со статистикой
        :return: None
        """
        env = Environment(loader=FileSystemLoader('./templates'))
        template = env.get_template("pdf_template.html")
        stats = []
        for year in self.salary_by_year.keys():
            stats.append([year, self.salary_by_year[year], self.vacs_per_year[year], self.salary_by_vac[year],
                          self.count_by_vac[year]])

        for key in self.city_percents:
            self.city_percents[key] = round(self.city_percents[key] * 100, 2)

        pdf_template = template \
            .render({'name': self.vacancy_name, 'path': '{0}/{1}'
                    .format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
                     'stats': stats, 'stats2': self.salary_by_city, 'stats3': self.city_percents})

        config = pdfkit.configuration(wkhtmltopdf=r'D:/Data1/wkhtmltopdf/bin/wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


if __name__ == '__main__':
    InputConnect()
